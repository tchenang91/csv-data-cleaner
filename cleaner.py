#!/usr/bin/env python3
"""CSV-Daten bereinigen und als formatierte XLSX-Datei exportieren.

Funktionen:
- Liest eine oder mehrere CSV-Dateien.
- Trennt Komma-separierte Inhalte in einzelne Spalten (auch wenn alles in einer Spalte steht).
- Entfernt Zeilen mit leerer `user_id`.
- Sortiert Datensätze alphabetisch nach `user_id`.
- Formatiert Werte:
  * `user_id` komplett UPPERCASE
  * andere Textwerte mit erstem Buchstaben in Großschreibung
- Exportiert nach "Erzeugte Datei.xlsx" mit einem Arbeitsblatt "Erzeugte Datei".
- Zellformatierung:
  * Alle Zellen zentriert
  * Alle Zellen mit Rahmenlinien
  * Kopfzeile grau + fett
  * Spalten `price` und `final_price` im Buchhaltungszahlenformat
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "Erzeugte Datei.xlsx"
SHEET_NAME = "Erzeugte Datei"
ACCOUNTING_FORMAT = '#,##0.00_);[Red](#,##0.00)'


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="CSV-Dateien bereinigen und in eine XLSX-Datei exportieren."
    )
    parser.add_argument(
        "inputs",
        nargs="+",
        type=Path,
        help="Pfad(e) zu CSV-Dateien (eine oder mehrere)",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path(OUTPUT_FILE),
        help=f"Ausgabedatei (Standard: {OUTPUT_FILE})",
    )
    return parser.parse_args()


def read_csv_rows(csv_path: Path) -> list[list[str]]:
    rows: list[list[str]] = []

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            if not row:
                continue

            # Falls die komplette Zeile als ein String vorliegt, erneut per CSV parsen.
            if len(row) == 1 and "," in row[0]:
                reparsed = next(csv.reader([row[0]]))
                rows.append([cell.strip() for cell in reparsed])
            else:
                rows.append([cell.strip() for cell in row])

    return rows


def find_user_id_index(headers: list[str]) -> int:
    normalized = [h.strip().lower().replace(" ", "") for h in headers]
    for i, h in enumerate(normalized):
        if h in {"userid", "user_id", "user-id"}:
            return i
    raise ValueError("Spalte 'user_id' wurde in den Headern nicht gefunden.")


def find_column_index(headers: list[str], column_name: str) -> int | None:
    target = column_name.strip().lower().replace(" ", "")
    normalized = [h.strip().lower().replace(" ", "") for h in headers]
    for i, h in enumerate(normalized):
        if h == target:
            return i
    return None


def title_case_like(value: str) -> str:
    value = value.strip()
    if not value:
        return value
    return value[0].upper() + value[1:].lower()


def to_number_maybe(value: str) -> float | str:
    v = value.strip().replace(".", "", 1).replace(",", "", 1)
    if not v:
        return ""
    try:
        return float(value)
    except ValueError:
        return value


def merge_and_clean(csv_paths: Iterable[Path]) -> tuple[list[str], list[list[str | float]]]:
    header: list[str] | None = None
    all_rows: list[list[str]] = []

    for csv_path in csv_paths:
        if not csv_path.exists():
            raise FileNotFoundError(f"Datei nicht gefunden: {csv_path}")

        rows = read_csv_rows(csv_path)
        if not rows:
            continue

        file_header, *data_rows = rows

        if header is None:
            header = file_header
        elif len(file_header) != len(header):
            raise ValueError(
                f"Inkompatible Headerlänge in {csv_path}: {len(file_header)} statt {len(header)}"
            )

        all_rows.extend(data_rows)

    if header is None:
        raise ValueError("Keine Daten gefunden. Bitte CSV-Dateien prüfen.")

    user_id_idx = find_user_id_index(header)

    cleaned: list[list[str | float]] = []
    for row in all_rows:
        if len(row) < len(header):
            row = row + [""] * (len(header) - len(row))
        elif len(row) > len(header):
            row = row[: len(header)]

        user_id = row[user_id_idx].strip()
        if not user_id:
            continue

        normalized_row: list[str | float] = []
        for idx, value in enumerate(row):
            value = value.strip()
            if idx == user_id_idx:
                normalized_row.append(value.upper())
            else:
                normalized_row.append(title_case_like(value) if value else "")
        cleaned.append(normalized_row)

    cleaned.sort(key=lambda r: str(r[user_id_idx]))

    # Preisfelder numerisch konvertieren
    price_idx = find_column_index(header, "price")
    final_price_idx = find_column_index(header, "final_price")

    for row in cleaned:
        if price_idx is not None:
            row[price_idx] = to_number_maybe(str(row[price_idx]))
        if final_price_idx is not None:
            row[final_price_idx] = to_number_maybe(str(row[final_price_idx]))

    return header, cleaned


def export_xlsx(header: list[str], rows: list[list[str | float]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(header)
    for row in rows:
        ws.append(row)

    # Styles
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
    header_font = Font(bold=True)

    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            cell.border = border
            if r == 1:
                cell.fill = header_fill
                cell.font = header_font

    # Buchhaltungsformat für price/final_price
    price_idx = find_column_index(header, "price")
    final_price_idx = find_column_index(header, "final_price")
    for col_idx in (price_idx, final_price_idx):
        if col_idx is None:
            continue
        excel_col = col_idx + 1
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=excel_col).number_format = ACCOUNTING_FORMAT

    # Optionale automatische Spaltenbreite
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        longest = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, max_row + 1))
        ws.column_dimensions[letter].width = min(max(longest + 2, 12), 45)

    wb.save(output_path)


def main() -> None:
    args = parse_args()
    header, rows = merge_and_clean(args.inputs)
    export_xlsx(header, rows, args.output)
    print(f"Fertig: {args.output}")


if __name__ == "__main__":
    main()
